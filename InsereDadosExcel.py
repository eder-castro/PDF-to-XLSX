import pandas as pd
from datetime import datetime
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math # Importar para usar math.isinf e math.isnan

# Nome do arquivo Excel
NOME_ARQUIVO_EXCEL = "CONTROLE FLUXO ORIGINAL.xlsx"
SHEET_NAME = "#NFs#"
HEADER_ROW_EXCEL = 2 # A linha do cabeçalho no Excel (Linha 2)
DATA_START_ROW_EXCEL = 3 # A linha onde os dados começam no Excel (Linha 3)

# Seus dicionários de novos dados
lista_de_dicionarios = [{'Numero_Nota': '36501', 'Data_Emissao': '01/08/2023', 'CNPJ_Prestador': '07601279000108', 'CNPJ_Tomador': '01666851868', 'Pedido': '', 'Contrato': '', 'Valor_Total': '900,00', 'Nome_Arquivo': '08 - ANGRA DOS REIS - NF 36501.pdf'},
{'Numero_Nota': '1521', 'Data_Emissao': '19/05/2025', 'CNPJ_Prestador': '12599272000139', 'CNPJ_Tomador': '13718634000126', 'Pedido': '4500012167', 'Contrato': '4700000307', 'Valor_Total': '548,33', 'Nome_Arquivo': '1521 Icon.pdf'},
{'Numero_Nota': '1735', 'Data_Emissao': '20/05/2025', 'CNPJ_Prestador': '59291534000167', 'CNPJ_Tomador': '04553060000192', 'Pedido': '4500012196', 'Contrato': '4700000341', 'Valor_Total': '2.590,09', 'Nome_Arquivo': '1735 NF - ICON - VSERRA.pdf'}]

# Variáveis para controle
novos_valores_a_adicionar = [] # Lista para armazenar dicionários já formatados para a planilha
qt_arquivos_plan = 0

# --- Mapeamento das colunas do dicionário para os nomes e índices das colunas da planilha Excel ---
colunas_para_inserir_info = {
    'NF': {'dict_key': 'Numero_Nota', 'type': int},
    'Data NF': {'dict_key': 'Data_Emissao', 'type': datetime},
    'CNPJ': {'dict_key': 'CNPJ_Prestador', 'type': str},
    'CNPJ/CPF Tomador': {'dict_key': 'CNPJ_Tomador', 'type': str},
    'Contrato': {'dict_key': 'Contrato', 'type': str},
    'Pedido': {'dict_key': 'Pedido', 'type': str},
    'Valor NF': {'dict_key': 'Valor_Total', 'type': float},
    'Nome do Arquivo': {'dict_key': 'Nome_Arquivo', 'type': str}
}

# --- Funções Auxiliares ---
def parse_data_emissao(data_str):
    """Tenta fazer o parse de uma string de data para um objeto datetime."""
    if not isinstance(data_str, str) or not data_str:
        return None
    try:
        dt = parser.parse(data_str)
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return dt
    except Exception:
        try:
            return datetime.strptime(data_str, "%d/%m/%Y")
        except Exception:
            return None

def formatar_valor_para_planilha(excel_col_name, value, expected_type):
    """
    Formata o valor para o tipo esperado pela coluna da planilha.
    Lida com NaN, None, strings vazias e floats infinitos/NaNs para evitar erros.
    """
    # 1. Tratar valores que são explicitamente None, NaN do pandas, ou strings vazias
    if pd.isna(value) or value is None or (isinstance(value, str) and not value.strip()):
        return None

    # 2. Tratar floats especiais (infinito, NaN) que podem não ser pegos por pd.isna() em alguns contextos
    if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
        return None

    if expected_type == datetime:
        return parse_data_emissao(value)
    elif expected_type == float:
        try:
            # Substitui vírgula por ponto para float e tenta converter
            return float(str(value).replace(",", "."))
        except (ValueError, TypeError):
            return None # Retorna None se não puder converter para float
    elif expected_type == int:
        try:
            # Converte para int, lidando com floats (ex: 123.0 -> 123)
            # É importante garantir que o valor já não seja inf/nan antes desta linha
            return int(float(value))
        except (ValueError, TypeError):
            return None # Retorna None se não puder converter para int
    elif expected_type == str:
        s_value = str(value)
        # Se for um número que veio com '.0' (ex: '12345678901234.0'), remove o '.0'
        if s_value.endswith('.0') and s_value[:-2].isdigit(): # Garante que é um número antes de remover
            s_value = s_value[:-2]
        return s_value.strip()
    else:
        return str(value).strip()


# --- Lógica Principal ---
try:
    # 1. Leitura de dados existentes para checagem de duplicatas usando pandas (otimizado)
    cols_to_read = ['NF', 'CNPJ']
    dados_existentes_para_checar = set()

    try:
        # Pandas é 0-indexed para header e skiprows.
        # HEADER_ROW_EXCEL=2 (Linha 2) -> header=1 para pandas
        # DATA_START_ROW_EXCEL=3 (Linha 3) -> skip a linha 1 (índice 0)
        df_existente = pd.read_excel(
            NOME_ARQUIVO_EXCEL,
            sheet_name=SHEET_NAME,
            header=HEADER_ROW_EXCEL - 1,
            usecols=cols_to_read,
            skiprows=[0] # Ignora a linha 1 oculta do Excel (0-indexed)
        )
        
        # Garante que as colunas são formatadas para comparação consistente
        # Aplica a função de formatação robusta
        df_existente['NF_formatted'] = df_existente['NF'].apply(lambda x: formatar_valor_para_planilha('NF', x, int))
        df_existente['CNPJ_formatted'] = df_existente['CNPJ'].apply(lambda x: formatar_valor_para_planilha('CNPJ', x, str))

        # Cria um set de tuplas (NF, CNPJ) para checagem rápida de duplicatas
        # Remove linhas onde NF ou CNPJ sejam None/NaN após a formatação
        dados_existentes_para_checar = set(
            tuple(row) for row in df_existente[['NF_formatted', 'CNPJ_formatted']].dropna().itertuples(index=False)
        )
        print(f"Total de NFs/CNPJs existentes na planilha (otimizado): {len(dados_existentes_para_checar)}")

    except Exception as e:
        print(f"Não foi possível ler as colunas 'NF' e 'CNPJ' com pandas para checagem de duplicatas. Erro: {e}")
        print("Continuando o processo com leitura de dados existente via openpyxl (pode ser mais lento para 20.000 linhas).")
        
        # Fallback para a checagem lenta se a leitura otimizada falhar
        book_fallback = load_workbook(NOME_ARQUIVO_EXCEL)
        sheet_fallback = book_fallback[SHEET_NAME]
        excel_headers_fallback = [cell.value for cell in sheet_fallback[HEADER_ROW_EXCEL]]
        
        try:
            idx_nf_fallback = excel_headers_fallback.index('NF') + 1
            idx_cnpj_fallback = excel_headers_fallback.index('CNPJ') + 1
        except ValueError:
            print("Erro crítico: Colunas 'NF' ou 'CNPJ' não encontradas nos cabeçalhos mesmo no fallback.")
            exit()

        for row_num in range(DATA_START_ROW_EXCEL, sheet_fallback.max_row + 1):
            nf_raw_value = sheet_fallback.cell(row=row_num, column=idx_nf_fallback).value
            cnpj_raw_value = sheet_fallback.cell(row=row_num, column=idx_cnpj_fallback).value
            
            # Formata os valores lidos da planilha para o tipo esperado para comparação
            nf_formatted = formatar_valor_para_planilha('NF', nf_raw_value, int)
            cnpj_formatted = formatar_valor_para_planilha('CNPJ', cnpj_raw_value, str)
            
            if nf_formatted is not None and cnpj_formatted is not None:
                dados_existentes_para_checar.add((nf_formatted, cnpj_formatted))
        print(f"Fallback: Total de NFs/CNPJs existentes (checado célula a célula): {len(dados_existentes_para_checar)}")
        
    # 2. Carregar o workbook com openpyxl para escrita (mantendo fórmulas, formatação, etc.)
    # Isso é feito aqui para garantir que o workbook esteja carregado para a etapa de escrita
    book = load_workbook(NOME_ARQUIVO_EXCEL)
    sheet = book[SHEET_NAME]

    # Obter os cabeçalhos da linha 2 para mapear as colunas
    excel_headers = [cell.value for cell in sheet[HEADER_ROW_EXCEL]]
    
    # 3. Encontrar a próxima linha para inserção
    # Manter a lógica de busca pela última linha preenchida para o "intervalo"
    # Este loop pode ser lento, mas só é executado uma vez.
    ultima_linha_preenchida = 0
    # Obter a lista de nomes das colunas que você preenche
    colunas_a_verificar_dados = [excel_col_name for excel_col_name in colunas_para_inserir_info.keys()]

    # Mapear os nomes das colunas para seus índices no Excel
    col_indices_para_verificar = []
    for col_name_check in colunas_a_verificar_dados:
        try:
            col_indices_para_verificar.append(excel_headers.index(col_name_check) + 1)
        except ValueError:
            # Se a coluna não for encontrada nos cabeçalhos, ignora-a para a verificação de "linha preenchida"
            pass 

    for r_idx in range(DATA_START_ROW_EXCEL, sheet.max_row + 1): # Começa a procurar a partir da linha 3
        has_data = False
        for col_idx_check in col_indices_para_verificar:
            cell_value = sheet.cell(row=r_idx, column=col_idx_check).value
            # Considera a célula como tendo dados se não for None/NaN e não for uma string vazia
            if cell_value is not None and not (isinstance(cell_value, str) and not cell_value.strip()) \
               and not (isinstance(cell_value, float) and (math.isinf(cell_value) or math.isnan(cell_value))):
                has_data = True
                break # Encontrou dados em alguma coluna, move para a próxima linha
        if has_data:
            ultima_linha_preenchida = r_idx
    
    next_insert_row = max(DATA_START_ROW_EXCEL, ultima_linha_preenchida + 1)
    
    print(f"Próxima linha para inserção de novos dados (determinada): {next_insert_row}")

    # 4. Processar novos dicionários e adicionar à lista de inserção
    for dicionario in lista_de_dicionarios:
        if dicionario:
            num_nota_formatado = formatar_valor_para_planilha(
                'NF',
                dicionario.get(colunas_para_inserir_info['NF']['dict_key']),
                colunas_para_inserir_info['NF']['type']
            )
            cnpj_prestador_formatado = formatar_valor_para_planilha(
                'CNPJ',
                dicionario.get(colunas_para_inserir_info['CNPJ']['dict_key']),
                colunas_para_inserir_info['CNPJ']['type']
            )

            # Só checa se os valores formatados não são None
            if num_nota_formatado is not None and cnpj_prestador_formatado is not None:
                existe = (num_nota_formatado, cnpj_prestador_formatado) in dados_existentes_para_checar
            else: # Se um dos valores essenciais for None, considera que "não existe" para inclusão
                existe = False
                print(f"[AVISO] Dados incompletos para NF/CNPJ no dicionário, impossível verificar duplicidade: {dicionario}")
            
            if not existe:
                linha_formatada = {}
                for excel_col_name, col_info in colunas_para_inserir_info.items():
                    dict_key = col_info['dict_key']
                    expected_type = col_info['type']
                    linha_formatada[excel_col_name] = formatar_valor_para_planilha(
                        excel_col_name, dicionario.get(dict_key), expected_type
                    )
                
                novos_valores_a_adicionar.append(linha_formatada)
                print("NF Incluída para processamento: ", dicionario)
                qt_arquivos_plan += 1
            else:
                print(f"[AVISO] NF {dicionario['Numero_Nota']} do CNPJ {dicionario['CNPJ_Prestador']} já existe na planilha.")
        else:
            print(f"Erro na leitura de um dicionário (dicionário vazio ou None).")

    # 5. Inserir os novos dados na planilha com openpyxl
    if novos_valores_a_adicionar:
        print(f"Total de novas NFs a serem adicionadas: {len(novos_valores_a_adicionar)}")

        col_indices = {}
        for excel_col_name in colunas_para_inserir_info.keys():
            try:
                col_indices[excel_col_name] = excel_headers.index(excel_col_name) + 1
            except ValueError:
                print(f"[AVISO] Coluna '{excel_col_name}' não encontrada nos cabeçalhos da planilha. Dados para esta coluna não serão inseridos.")
                col_indices[excel_col_name] = None

        for new_row_data in novos_valores_a_adicionar:
            for excel_col_name, value in new_row_data.items():
                col_idx = col_indices.get(excel_col_name)
                if col_idx is not None:
                    cell = sheet.cell(row=next_insert_row, column=col_idx)
                    cell.value = value
            next_insert_row += 1
        
        book.save(NOME_ARQUIVO_EXCEL)
        print(f"Novas NFs adicionadas com sucesso à planilha {NOME_ARQUIVO_EXCEL}.")
    else:
        print("Nenhuma nova NF para adicionar.")

except FileNotFoundError:
    print(f"Erro: O arquivo '{NOME_ARQUIVO_EXCEL}' não foi encontrado. Certifique-se de que ele existe e está no diretório correto.")
except Exception as e:
    print(f"Ocorreu um erro inesperado: {e}")

print(f"Total de NFs processadas (incluídas ou verificadas): {qt_arquivos_plan}")